from openpyxl import load_workbook
from datetime import datetime

file_path = 'prova.xlsx'

wb = load_workbook(file_path)
ws = wb['Hoja1']  # or wb.active
ws['G6'] = datetime.now()
ws["G6"].number_format = "dd/mm/yyyy"
#lax = wb.defined_names['laequis']

wb.save(file_path)
